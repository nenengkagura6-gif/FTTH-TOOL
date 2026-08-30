"""
Pastikan template default_apd.xlsx tidak lagi membawa data cluster lama,
dan default template yang sah tetap utuh.

Kenapa penting: kalau engine gagal mengisi sebuah sel — koordinat FDT tidak
ketemu, atau Nominatim tidak mengembalikan kecamatan — nilai lama akan
tertinggal di berkas hasil dan terbaca seolah data asli cluster yang baru.
Itulah yang terjadi pada laporan "tikor masih template" dan "kecamatan kosong".
"""
import os
import sys

from openpyxl import load_workbook

TEMPLATE = os.path.join(os.path.dirname(__file__), "..", "app", "templates",
                        "default_apd.xlsx")

failures = 0


def check(name, cond, detail=""):
    global failures
    if cond:
        print("  [OK]    %s %s" % (name, detail))
    else:
        print("  [GAGAL] %s %s" % (name, detail))
        failures += 1


wb = load_workbook(TEMPLATE)
ws = wb.worksheets[0]

print("=== 1. Tidak ada sisa data cluster lama ===")

# Nama-nama milik cluster contoh yang dulu tertinggal di template
KOTOR = ["SITIWANGUN", "SITIWINANGUN", "JAMBLANG", "BAKUNG", "45156"]

temuan = []
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if isinstance(v, str):
            u = v.upper()
            for k in KOTOR:
                if k in u:
                    temuan.append((cell.coordinate, v))
                    break

check("tidak ada nama cluster lama tersisa", not temuan,
      "-> %s" % (temuan[:5] if temuan else ""))

for sel, label in [("C3", "Region"), ("C5", "Cluster Name"), ("N3", "FDT TYPE")]:
    check("%s (%s) kosong" % (sel, label), ws[sel].value in (None, ""),
          "-> %r" % (ws[sel].value,))

check("N6 (COORDINATE FDT) tanpa koordinat lama",
      str(ws["N6"].value or "").strip() in (":", ": ", ""),
      "-> %r" % (ws["N6"].value,))

for sel, label in [("N10", "regency_city"), ("O10", "district"),
                   ("P10", "subdistrict"), ("Q10", "postalcode"),
                   ("W10", "Area"), ("Y10", "Clustername"),
                   ("Z10", "Commercial_name"), ("AD10", "street")]:
    check("%s (%s) kosong" % (sel, label), ws[sel].value in (None, ""),
          "-> %r" % (ws[sel].value,))

m_terisi = sum(1 for r in range(10, ws.max_row + 1)
               if ws.cell(row=r, column=13).value not in (None, ""))
check("kolom M (Province) kosong seluruhnya", m_terisi == 0,
      "-> %d baris masih terisi" % m_terisi)


print()
print("=== 2. Default template yang sah TETAP ada ===")

DEFAULT = {
    "U": ("competition", "FTTH"),
    "AC": ("address_prefix", "JLN."),
    "AX": ("residential_service_ready", "T"),
    "AY": ("sme_service_ready", "F"),
    "AZ": ("enterprise_service_ready", "F"),
    "BA": ("installation", "AE"),
    "BB": ("availability", "T"),
    "BC": ("network_presence", 31),
    "BD": ("wallplate_installation", "F"),
}
for kolom, (label, nilai) in DEFAULT.items():
    v = ws[f"{kolom}10"].value
    check("%s (%s) = %r" % (kolom, label, nilai), str(v) == str(nilai), "-> %r" % (v,))

check("BE10 rumus nomor baris utuh",
      str(ws["BE10"].value or "").startswith("=ROW()"), "-> %r" % (ws["BE10"].value,))
check("BF10 rumus concatenate utuh",
      "CONCATENATE" in str(ws["BF10"].value or ""), "-> %r" % (ws["BF10"].value,))


print()
print("=== 3. Struktur & format tidak rusak ===")

check("label kolom baris 8 utuh",
      ws["A8"].value == "FDT Tray (Front)" and ws["O8"].value == "district"
      and ws["P8"].value == "subdistrict")
check("baris 'mandatory' utuh", ws["A9"].value == "mandatory")
check("autofilter masih ada", ws.auto_filter.ref is not None,
      "-> %s" % ws.auto_filter.ref)
check("lebar kolom terjaga", len(ws.column_dimensions) > 40,
      "-> %d kolom" % len(ws.column_dimensions))
check("judul sheet utuh", ws.title.startswith("Homepass Database"),
      "-> %r" % ws.title)
check("rumus N4 (jumlah HP) utuh", "COUNTA" in str(ws["N4"].value or ""),
      "-> %r" % (ws["N4"].value,))

print()
print("HASIL:", "SEMUA LULUS" if failures == 0 else "%d KEGAGALAN" % failures)
sys.exit(1 if failures else 0)

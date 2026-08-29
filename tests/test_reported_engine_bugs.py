"""Bukti perbaikan tiga bug engine yang dilaporkan."""
import sys
import io as _io

sys.path.insert(0, r"D:\WEB APP\app")

failures = 0


def check(name, cond, detail=""):
    global failures
    if cond:
        print("  [OK]    %s %s" % (name, detail))
    else:
        print("  [GAGAL] %s %s" % (name, detail))
        failures += 1


# =====================================================================
# BUG 1a — koordinat FDT pada KML BERNAMESPACE
# =====================================================================
print("=== BUG 1a: koordinat FDT di KML bernamespace ===")

KML_NS = b"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document><name>Cluster</name>
    <Folder><name>FDT</name>
      <Folder><name>Titik</name>
        <Placemark><name>FDT 01</name>
          <Point><coordinates>106.83000,-6.19600,0</coordinates></Point>
        </Placemark>
      </Folder>
      <Placemark><name>FDT 02</name>
        <Point><coordinates>106.84000,-6.20600,0</coordinates></Point>
      </Placemark>
    </Folder>
  </Document>
</kml>"""

from engines.apd_engine import APDEngine

eng = APDEngine()
eng.load_kml(KML_NS, "uji.kml", is_kmz=False)
coords = eng.get_all_fdt_coords()

check("FDT 01 terbaca", "FDT 01" in coords, "-> %s" % (coords.get("FDT 01"),))
check("FDT 02 terbaca (dalam sub-folder / anak langsung)", "FDT 02" in coords,
      "-> %s" % (coords.get("FDT 02"),))
check("koordinat benar", coords.get("FDT 01") == ("-6.19600", "106.83000"),
      "-> %s" % (coords.get("FDT 01"),))


# =====================================================================
# BUG 1b — kecamatan tidak lagi terserobot kabupaten
# =====================================================================
print()
print("=== BUG 1b: pemetaan kecamatan ===")


class FakeResp:
    status_code = 200

    def __init__(self, addr):
        self._addr = addr

    def json(self):
        return {"address": self._addr}


class FakeSession:
    def __init__(self, addr):
        self.addr = addr

    def get(self, *a, **k):
        return FakeResp(self.addr)


import time as _time
_time.sleep = lambda *a, **k: None  # jangan tunggu rate limit saat uji

# Kasus khas Indonesia: kecamatan ada di field "municipality"
eng2 = APDEngine()
eng2.session = FakeSession({
    "state": "Jawa Barat",
    "county": "Kabupaten Bandung",
    "municipality": "Kecamatan Soreang",
    "village": "Desa Pamekaran",
    "postcode": "40911",
    "road": "Jalan Raya Soreang",
})
res = eng2.reverse_geocode(-7.0, 107.5)
check("kecamatan terisi dari 'municipality'", res["kecamatan"] != "", "-> %r" % res["kecamatan"])
check("kabupaten tidak terserobot", "BANDUNG" in res["kabupaten"], "-> %r" % res["kabupaten"])
check("desa terpisah dari kecamatan", res["desa"] != res["kecamatan"],
      "-> desa=%r kecamatan=%r" % (res["desa"], res["kecamatan"]))


# =====================================================================
# BUG 2 — BOQ satu FDT tidak boleh menyalin ke kolom FDT 02 / 03
# =====================================================================
print()
print("=== BUG 2: BOQ satu FDT ===")

KML_BOQ = b"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document><name>Cluster</name>
    <Folder><name>LINE A FDT 01</name>
      <Folder><name>FAT</name>
        <Placemark><name>FAT-1</name><Point><coordinates>106.8,-6.2,0</coordinates></Point></Placemark>
        <Placemark><name>FAT-2</name><Point><coordinates>106.8,-6.2,0</coordinates></Point></Placemark>
      </Folder>
      <Folder><name>NEW POLE 7-4</name>
        <Placemark><name>P1</name><Point><coordinates>106.8,-6.2,0</coordinates></Point></Placemark>
        <Placemark><name>P2</name><Point><coordinates>106.8,-6.2,0</coordinates></Point></Placemark>
        <Placemark><name>P3</name><Point><coordinates>106.8,-6.2,0</coordinates></Point></Placemark>
      </Folder>
    </Folder>
  </Document>
</kml>"""

from engines.kml_engine import KMLEngine

boq = KMLEngine()
boq.load_kml(KML_BOQ, "uji_boq.kml", is_kmz=False)
out = boq.process()
check("proses BOQ sukses", out.get("status") == "success", "-> %s" % out.get("status"))

from openpyxl import load_workbook
wb = load_workbook(_io.BytesIO(out["content"]))
ws = wb["BoM AE"] if "BoM AE" in wb.sheetnames else wb.active

# FAT LINE A -> baris 36 ; NEW POLE 7-4 -> baris 54
fat_c, fat_i, fat_o = ws["C36"].value, ws["I36"].value, ws["O36"].value
pol_c, pol_i, pol_o = ws["C54"].value, ws["I54"].value, ws["O54"].value

check("FAT masuk kolom C (FDT 01)", fat_c == 2, "-> C36=%r" % fat_c)
check("FAT TIDAK bocor ke kolom I (FDT 02)", not fat_i, "-> I36=%r" % fat_i)
check("FAT TIDAK bocor ke kolom O (FDT 03)", not fat_o, "-> O36=%r" % fat_o)
check("Tiang masuk kolom C (FDT 01)", pol_c == 3, "-> C54=%r" % pol_c)
check("Tiang TIDAK bocor ke kolom I (FDT 02)", not pol_i, "-> I54=%r" % pol_i)
check("Tiang TIDAK bocor ke kolom O (FDT 03)", not pol_o, "-> O54=%r" % pol_o)

print()
print("HASIL:", "SEMUA LULUS" if failures == 0 else "%d KEGAGALAN" % failures)
sys.exit(1 if failures else 0)

from lxml import etree as ET
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
import math
import requests
import time
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import glob
import os
import sys

# ==== helper aman ambil nama tag ====
def safe_localname(elem):
    try:
        return ET.QName(elem.tag).localname
    except Exception:
        return None

# ==== fungsi hitung jarak (Haversine) ====
def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

# ==== parsing KML untuk FAT & POLE ====
def parse_fat_pole(filename):
    parser = ET.XMLParser(recover=True)
    tree = ET.parse(filename, parser)
    root = tree.getroot()

    placemarks_fat = []
    placemarks_pole = []

    for folder in root.iter():
        if safe_localname(folder) == "Folder":
            folder_name = ""
            for child in folder:
                if safe_localname(child) == "name":
                    folder_name = child.text if child.text else ""
                    break
            for pm in folder:
                if safe_localname(pm) == "Placemark":
                    name = "NONAME"
                    coords = None
                    for child in pm.iter():
                        if safe_localname(child) == "name":
                            name = child.text if child.text else "NONAME"
                        if safe_localname(child) == "coordinates":
                            coords = child
                    if coords is not None:
                        coords_text = coords.text.strip()
                        for coord_pair in coords_text.split():
                            try:
                                lon, lat, *_ = map(float, coord_pair.split(","))
                                if folder_name.strip().upper() == "FAT":
                                    match = re.search(r'\b([A-Z]\d{1,2})\b', name, re.IGNORECASE)
                                    fat_id = match.group(1).upper() if match else name.strip().upper()
                                    placemarks_fat.append({"fat_id": fat_id, "lat": lat, "lon": lon})
                                elif "POLE" in folder_name.upper():
                                    placemarks_pole.append({"name": name.strip(), "lat": lat, "lon": lon})
                                break
                            except:
                                continue
    return placemarks_fat, placemarks_pole

# ==== cari pasangan FAT ↔ POLE ====
def build_fat_to_pole_map(fats, poles, max_distance=15):
    fat_to_pole = {}
    for fat in fats:
        fat_id, flat, flon = fat["fat_id"], fat["lat"], fat["lon"]
        best_match, best_distance = None, float("inf")
        for pole in poles:
            d = haversine(flat, flon, pole["lat"], pole["lon"])
            if d < best_distance:
                best_distance = d
                best_match = pole
        if best_match and best_distance <= max_distance:
            fat_to_pole[fat_id] = best_match
    return fat_to_pole

# ==== cari pasangan FAT ↔ HP COVER (≤25m) ====
def build_fat_to_hpc_map(fats, hp_points, max_distance=25):
    fat_to_hpc = {}
    for fat in fats:
        fat_id, flat, flon = fat["fat_id"], fat["lat"], fat["lon"]
        best_match, best_distance = None, float("inf")
        for hp in hp_points:
            d = haversine(flat, flon, float(hp[1]), float(hp[2]))
            if d < best_distance:
                best_distance = d
                best_match = hp
        if best_match and best_distance <= max_distance:
            fat_to_hpc[fat_id] = best_match
    return fat_to_hpc

# ==== HP COVER ====
def find_hp_cover_folders(root_element):
    hp_cover_folders = []
    for folder in root_element.iter():
        if safe_localname(folder) == "Folder":
            name_elem = None
            for child in folder:
                if safe_localname(child) == "name":
                    name_elem = child
                    break
            if name_elem is not None and name_elem.text and "HP" in name_elem.text.upper() and "COVER" in name_elem.text.upper():
                hp_cover_folders.append(folder)
    return hp_cover_folders

def extract_points_recursive(element, current_folder):
    points = []
    for child in element:
        tag = safe_localname(child)
        if tag == "Folder":
            name_elem = None
            for sub in child:
                if safe_localname(sub) == "name":
                    name_elem = sub
                    break
            subfolder_name = name_elem.text.strip() if name_elem is not None else current_folder
            new_folder = f"{current_folder}/{subfolder_name}" if current_folder else subfolder_name
            points.extend(extract_points_recursive(child, new_folder))
        elif tag == "Placemark":
            name = "No Name"
            lon, lat = "", ""
            for sub in child.iter():
                if safe_localname(sub) == "name":
                    name = sub.text.strip() if sub.text else "No Name"
                if safe_localname(sub) == "coordinates":
                    lon, lat = sub.text.strip().split(",")[:2]
            if lat and lon:
                points.append([name, lat, lon, current_folder])
    return points


# =====================================
# NOMINATIM REVERSE GEOCODE
# =====================================

session = requests.Session()

retry = Retry(
    total=3,
    backoff_factor=1,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET"]
)

adapter = HTTPAdapter(max_retries=retry)

session.mount("http://", adapter)
session.mount("https://", adapter)

GEOCODE_CACHE = {}

def clean_region_name(text):

    if not text:
        return ""

    text = str(text)

    remove_words = [
        "Provinsi ",
        "Kecamatan ",
        "Kabupaten ",
        "Kab. ",
        "Kota ",
        "Desa ",
        "Kelurahan ",
    ]

    for rw in remove_words:
        text = text.replace(rw, "")

    return text.strip().upper()



def reverse_geocode(lat, lon):

    try:

        key = f"{round(float(lat), 6)},{round(float(lon), 6)}"

        if key in GEOCODE_CACHE:
            return GEOCODE_CACHE[key]

        url = "https://nominatim.openstreetmap.org/reverse"

        headers = {
            "User-Agent": "Mozilla/5.0 APD_HPDB_SCRIPT"
        }

        params = {
            "lat": lat,
            "lon": lon,
            "format": "jsonv2",
            "addressdetails": 1,
            "zoom": 18
        }

        response = session.get(
            url,
            params=params,
            headers=headers,
            timeout=30
        )

        time.sleep(1.2)

        if response.status_code != 200:

            return {
                "province": "",
                "kabupaten": "",
                "kecamatan": "",
                "desa": "",
                "kodepos": ""
            }

        data = response.json()

        addr = data.get("address", {})

        province = clean_region_name(
            addr.get("state", "")
        )

        kabupaten = clean_region_name(
            addr.get("city", "")
            or addr.get("county", "")
            or addr.get("municipality", "")
            or addr.get("state_district", "")
        )

        kecamatan = clean_region_name(
    addr.get("subdistrict", "")
    or addr.get("district", "")
    or addr.get("city_district", "")
    or addr.get("suburb", "")
    or addr.get("borough", "")
    or addr.get("quarter", "")
    or addr.get("neighbourhood", "")
        )

        desa = clean_region_name(
            addr.get("village", "")
            or addr.get("hamlet", "")
            or addr.get("neighbourhood", "")
            or addr.get("quarter", "")
        )

        kodepos = str(addr.get("postcode", "")).strip()

        result = {
            "province": province,
            "kabupaten": kabupaten,
            "kecamatan": kecamatan,
            "desa": desa,
            "kodepos": kodepos
        }

        GEOCODE_CACHE[key] = result

        return result

    except Exception:

        return {
            "province": "",
            "kabupaten": "",
            "kecamatan": "",
            "desa": "",
            "kodepos": ""
        }


# ==== MAIN ====
kml_files = glob.glob("*.kml")
if not kml_files:
    print("❌ Tidak ada file .kml ditemukan di folder ini!")
    sys.exit(1)

input_kml = kml_files[0]
output_xlsx = os.path.splitext(input_kml)[0] + "_with_pole.xlsx"

print(f"✅ File KML terdeteksi: {input_kml}")

parser = ET.XMLParser(recover=True)
tree = ET.parse(input_kml, parser)
root = tree.getroot()

hp_folders = find_hp_cover_folders(root)
all_points = []
for folder in hp_folders:
    all_points.extend(extract_points_recursive(folder, ""))

fats, poles = parse_fat_pole(input_kml)
fat_to_pole = build_fat_to_pole_map(fats, poles, max_distance=15)
fat_to_hpc = build_fat_to_hpc_map(fats, all_points, max_distance=25)

fat_id_nums = {}
for row in all_points:
    match = re.search(r'\b([A-Z]\d{1,2})\b', row[3] + " " + row[0], re.IGNORECASE)
    fat_id = match.group(1).upper() if match else ""
    if not fat_id:
        continue
    prefix = fat_id[0].upper()
    num = int(re.findall(r'\d+', fat_id)[0]) if re.findall(r'\d+', fat_id) else 0
    if prefix not in fat_id_nums:
        fat_id_nums[prefix] = []
    fat_id_nums[prefix].append(num)
fat_id_max = {k: max(v) for k, v in fat_id_nums.items()}

wb = Workbook()
ws = wb.active
ws.append([
    "FDT Tray (Front)", "FDT Port", "Line", "Capacity",
    "Tube Colour", "Core Number", "FAT ID",
    "FAT Port", "POLE_Name", "POLE_Lat", "POLE_Lon",
    "HP_Cover",
    "name_1", "name_2", "latitude", "longitude"
])

fat_port_counters = {}
fdt_port_counter = 1
core_number = 1
skip_numbers = {11, 12, 23, 24, 35, 36, 47, 48}
last_fat_id = ""
last_prefix = ""

for row in all_points:
    name, lat, lon, folder_path = row
    name_parts = [p.strip() for p in re.split(r"[,/ ]", name) if p.strip()]
    name_1 = name_parts[0] if len(name_parts) > 0 else ""
    name_2 = name_parts[1] if len(name_parts) > 1 else ""

    match = re.search(r'\b([A-Z]\d{1,2})\b', folder_path + " " + name, re.IGNORECASE)
    fat_id = match.group(1).upper() if match else ""
    if not fat_id:
        continue

    prefix = fat_id[0].upper()
    num_part = int(re.findall(r'\d+', fat_id)[0]) if re.findall(r'\d+', fat_id) else 0

    if fat_id != last_fat_id:
        fat_port_counters[fat_id] = 1
        last_fat_id = fat_id

    if prefix != last_prefix:
        core_number = 1
        last_prefix = prefix

    fat_port = fat_port_counters[fat_id]
    fat_port_counters[fat_id] += 1

    if fat_port == 1:
        fdt_port = fdt_port_counter
        fdt_port_counter += 1
    else:
        fdt_port = ""

    core = ""
    if fat_port in [1, 2]:
        while core_number in skip_numbers:
            core_number += 1
        core = core_number
        core_number += 1

    line = ""
    cap = ""
    tube_number = ""
    if core:
        line = f"LINE {prefix}"
        max_num = fat_id_max.get(prefix, 0)
        if max_num <= 10:
            cap = "24C/2T"
        elif max_num <= 15:
            cap = "36C/3T"
        elif max_num <= 20:
            cap = "48C/4T"
        if num_part <= 5:
            tube_number = "1"
        elif num_part <= 10:
            tube_number = "2"
        elif num_part <= 15:
            tube_number = "3"
        elif num_part <= 20:
            tube_number = "4"

    if isinstance(fdt_port, int):
        if 1 <= fdt_port <= 8: tray = 1
        elif 9 <= fdt_port <= 16: tray = 2
        elif 17 <= fdt_port <= 24: tray = 3
        elif 25 <= fdt_port <= 32: tray = 4
        elif 33 <= fdt_port <= 40: tray = 5
        else: tray = ""
    else:
        tray = ""

    pole_name, pole_lat, pole_lon = "", "", ""
    if fat_id in fat_to_pole:
        pole_name = fat_to_pole[fat_id]["name"]
        pole_lat = fat_to_pole[fat_id]["lat"]
        pole_lon = fat_to_pole[fat_id]["lon"]

    hpc_text = ""
    if fat_id in fat_to_hpc:
        hpc_text = "IN FRONT OF HP NUMBER " + fat_to_hpc[fat_id][0]

    ws.append([
        tray, fdt_port, line, cap, tube_number, core, fat_id, fat_port,
        pole_name, pole_lat, pole_lon,
        hpc_text,
        name_1, name_2, lat, lon
    ])

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=2):
    for idx, cell in enumerate(row, 1):
        if cell.value not in (None, ""):
            cell.border = thin_border
        if idx == 5:
            if cell.value == "1":
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            elif cell.value == "2":
                cell.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
            elif cell.value == "3":
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif cell.value == "4":
                cell.fill = PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid")

try:
    wb.save(output_xlsx)
except PermissionError:
    alt_name = output_xlsx.replace(".xlsx", "_new.xlsx")
    print(f"⚠️ File sedang dibuka. Menyimpan dengan nama lain: {alt_name}")
    wb.save(alt_name)
    output_xlsx = alt_name

print(f"✔ File XLSX berhasil dibuat: {output_xlsx}")

# ==== Integrasi hasil ke Template Excel (versi fix warna dan posisi) ====
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

template_path = "APD_HPDB_.xlsx"
output_final = f"APD_HPDB_{os.path.basename(output_xlsx)}"

wb_template = load_workbook(template_path)
ws_template = wb_template.worksheets[0]

wb_data = load_workbook(output_xlsx)
ws_data = wb_data.active

headers = [str(h).strip() for h in next(ws_data.iter_rows(min_row=1, max_row=1, values_only=True))]
data_rows = list(ws_data.iter_rows(min_row=2, values_only=True))

header_map = {
    "FDT Tray (Front)": "A",
    "FDT Port": "B",
    "Line": "C",
    "Capacity": "D",
    "Tube Colour": "E",
    "Core Number": "F",
    "FAT ID": ["G", "AT"],
    "FAT Port": "H",
    "POLE_Name": "I",
    "POLE_Lat": "J",
    "POLE_Lon": "K",
    "HP_Cover": "L",
    "name_1": "AJ",
    "name_2": "AJ",  # gabung di belakang name_1
    "latitude": "AV",
    "longitude": "AW",
}

# warna berdasarkan nilai tube colour
tube_colors = {
    "1": "ADD8E6",  # biru muda
    "2": "FFD580",  # oranye muda
    "3": "90EE90",  # hijau muda
    "4": "D2B48C",  # coklat muda
}

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

start_row = 10
for i, row in enumerate(data_rows, start=start_row):
    row_dict = dict(zip(headers, row))

    for key, target in header_map.items():
        if key not in row_dict:
            continue
        val = row_dict[key]
        if val in (None, ""):
            continue

        if isinstance(target, list):  # FAT ID ke G dan AT
            for col in target:
                cell = ws_template[f"{col}{i}"]
                cell.value = val
                cell.border = thin_border
        elif key == "name_1":  # gabungkan dengan name_2 di AJ
            combined = str(val)
            if row_dict.get("name_2"):
                combined += " " + str(row_dict["name_2"])
            cell = ws_template[f"{target}{i}"]
            cell.value = combined
            cell.border = thin_border
        elif key == "name_2":
            continue
        else:
            cell = ws_template[f"{target}{i}"]
            if key in ["POLE_Lat", "POLE_Lon", "latitude", "longitude"]:
                try:
                    val = float(val)
                except:
                    pass

            cell.value = val
            cell.border = thin_border

            # warnai Tube Colour berdasarkan nilainya
            if key == "Tube Colour":
                color_hex = tube_colors.get(str(val).strip(), None)
                if color_hex:
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")


# ===============================
# ✨ FITUR TAMBAHAN: KOORDINAT FDT + NAMA FILE TANPA KODE
# ===============================
import xml.etree.ElementTree as ET
import os

def safe_localname(elem):
    if elem.tag[0] == "{":
        return elem.tag.split("}", 1)[1]
    return elem.tag

def get_fdt_coords(kml_file):
    try:
        parser = ET.XMLParser()
        tree = ET.parse(kml_file, parser)
        root = tree.getroot()
        for folder in root.iter():
            if safe_localname(folder) == "Folder":
                name_elem = None
                for child in folder:
                    if safe_localname(child) == "name":
                        name_elem = child
                        break
                if name_elem is not None and name_elem.text and name_elem.text.strip().upper() == "FDT":
                    for pm in folder:
                        if safe_localname(pm) == "Placemark":
                            for coord in pm.iter():
                                if safe_localname(coord) == "coordinates":
                                    lon, lat = map(float, coord.text.strip().split(",")[:2])
                                    return f"{lat:.5f}", f"{lon:.5f}"
    except Exception as e:
        print(f"Gagal ambil koordinat FDT: {e}")
    return None, None

# --- Ambil koordinat dari file input KML ---
lat_fdt, lon_fdt = get_fdt_coords(input_kml)
if lat_fdt and lon_fdt:
    ws_template["N6"].value = f": {lat_fdt},{lon_fdt}"

# --- Ambil nama file tanpa kode ---
file_base_name = os.path.splitext(os.path.basename(input_kml))[0]
file_base_name_clean = file_base_name.replace("APD_HPDB_", "")
parts = file_base_name_clean.split(" ", 1)
if len(parts) > 1:
    lokasi_nama = parts[1].strip()
else:
    lokasi_nama = file_base_name_clean.strip()

ws_template["C5"].value = lokasi_nama
ws_template["Y10"].value = lokasi_nama
ws_template["Z10"].value = lokasi_nama


# =====================================
# REVERSE GEOCODE
# =====================================

max_row_data = start_row + len(data_rows) - 1

lat_source = ws_template["J10"].value
lon_source = ws_template["K10"].value

if lat_source in (None, "") or lon_source in (None, ""):

    lat_source = ws_template["AV10"].value
    lon_source = ws_template["AW10"].value

geo_result = {
    "province": "",
    "kabupaten": "",
    "kecamatan": "",
    "desa": "",
    "kodepos": ""
}

if (
    lat_source not in (None, "")
    and lon_source not in (None, "")
):

    geo_result = reverse_geocode(
        lat_source,
        lon_source
    )

ws_template["M10"] = geo_result["province"]
ws_template["N10"] = geo_result["kabupaten"]
ws_template["W10"] = geo_result["kabupaten"]
ws_template["O10"] = geo_result["kecamatan"]
ws_template["P10"] = geo_result["desa"]
ws_template["Q10"] = geo_result["kodepos"]

for r in range(11, max_row_data + 1):

    ws_template[f"M{r}"] = ws_template["M10"].value
    ws_template[f"N{r}"] = ws_template["N10"].value
    ws_template[f"W{r}"] = ws_template["W10"].value
    ws_template[f"O{r}"] = ws_template["O10"].value
    ws_template[f"P{r}"] = ws_template["P10"].value
    ws_template[f"Q{r}"] = ws_template["Q10"].value

# =====================================
# FORMULA N4 OTOMATIS
# =====================================

ws_template["N4"] = (
    f'=\": \"&COUNTA(G10:G{max_row_data})&\" HP /Aerial\"'
)

# =====================================
# HAPUS ROW TIDAK TERPAKAI
# =====================================

last_template_row = ws_template.max_row

if max_row_data < last_template_row:

    delete_start = max_row_data + 1

    delete_count = (
        last_template_row - max_row_data
    )

    ws_template.delete_rows(
        delete_start,
        delete_count
    )


wb_template.save(output_final)
print(f"🎯 Data berhasil dimasukkan ke template dan disimpan sebagai: {output_final}")

# ===============================
# ✨ RENAME FILE EXCEL SESUAI FORMAT AKHIR (FIXED)
# ===============================
excel_path = output_final  # gunakan file hasil akhir sebagai path awal

excel_name = os.path.basename(excel_path)
excel_base = os.path.splitext(excel_name)[0]

# 1. Hilangkan akhiran _with_pole
excel_base = excel_base.replace("_with_pole", "")

# 2. Hilangkan prefix APD_HPDB_ sementara
prefix = "APD_HPDB_"
if excel_base.startswith(prefix):
    after_prefix = excel_base[len(prefix):]
else:
    after_prefix = excel_base

# 3. Hilangkan kode tengah seperti CBN005496
parts = after_prefix.split(" ", 1)
if len(parts) > 1:
    lokasi_nama_final = parts[1].strip()
else:
    lokasi_nama_final = after_prefix.strip()

# 4. Satukan kembali nama file dengan prefix APD_HPDB_
final_name = f"{prefix}{lokasi_nama_final}.xlsx"

# 5. Tentukan path final excel
final_excel_path = os.path.join(os.path.dirname(excel_path), final_name)
if excel_path != final_excel_path:
    if os.path.exists(final_excel_path):
        os.remove(final_excel_path)
    os.rename(excel_path, final_excel_path)

print(f"📁 File akhir disimpan sebagai: {final_excel_path}")

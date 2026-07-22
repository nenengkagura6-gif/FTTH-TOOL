"""
Engine for APD HPDB Processing
Supports both KML and KMZ files
"""
import io
import os
import re
import math
import time
import zipfile
import requests
from typing import Dict, List, Tuple, Any, Optional
from lxml import etree
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter

from utils.commons import haversine, safe_localname
from collections import defaultdict


def get_fdt_name_from_folder_path(folder_path: str) -> str:
    """Extract FDT name from folder path (e.g., 'LINE A FDT 02/HP COVER' -> 'FDT 02')"""
    match = re.search(r'\bFDT\s*(\d+)\b', folder_path, re.IGNORECASE)
    if match:
        return f"FDT {int(match.group(1)):02d}"
    return "FDT 01" # Default fallback


def parse_kml_lxml_with_kmz(content: bytes, is_kmz: bool = False) -> etree.ElementTree:
    """Parse KML content using lxml, with KMZ support."""
    parser = etree.XMLParser(resolve_entities=False, no_network=True, recover=True)
    
    if is_kmz:
        with zipfile.ZipFile(io.BytesIO(content), "r") as kmz:
            kml_files = [f for f in kmz.namelist() if f.endswith(".kml")]
            if not kml_files:
                raise ValueError("No KML file found inside KMZ archive")
            kml_name = kml_files[0]
            with kmz.open(kml_name) as kml_file:
                content = kml_file.read()
    
    return etree.parse(io.BytesIO(content), parser)


class APDEngine:
    """Engine for APD HPDB processing with geocoding."""
    
    def __init__(self, template_content: bytes = None, apd_template_content: bytes = None):
        self.template_content = template_content
        
        if apd_template_content:
            self.apd_template_content = apd_template_content
        else:
            base_dir = Path(__file__).resolve().parent.parent
            default_template = base_dir / "templates" / "default_apd.xlsx"
            
            if default_template.exists():
                with open(default_template, "rb") as f:
                    self.apd_template_content = f.read()
            else:
                self.apd_template_content = None
        
        self.session = self._create_session()
        self.geocode_cache = {}
        self.input_filename = ""
        self.tree = None
        self.root = None
        self.fats = []
        self.poles = []
        self.hp_points = []
    
    def _create_session(self) -> requests.Session:
        """Create requests session with retry logic."""
        session = requests.Session()
        retry = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"]
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        return session
    
    def load_kml(self, content: bytes, filename: str, is_kmz: bool = False) -> Dict[str, Any]:
        """Load KML or KMZ content."""
        self.input_filename = filename
        self.tree = parse_kml_lxml_with_kmz(content, is_kmz)
        self.root = self.tree.getroot()
        return {"status": "success", "filename": filename, "is_kmz": is_kmz}
    
    def set_templates(self, template_content: bytes = None, apd_template_content: bytes = None) -> None:
        """Set template contents."""
        self.template_content = template_content
        self.apd_template_content = apd_template_content
    
    def parse_fat_pole(self) -> Tuple[List[Dict], List[Dict]]:
        """Parse FAT and POLE placemarks from KML."""
        placemarks_fat = []
        placemarks_pole = []
        
        for folder in self.root.iter():
            if safe_localname(folder) == "Folder":
                folder_name = ""
                for child in folder:
                    if safe_localname(child) == "name":
                        folder_name = child.text if child.text else ""
                        break
                
                for pm in folder:
                    if safe_localname(pm) == "Placemark":
                        name = "NONAME"
                        coords = None
                        for child in pm.iter():
                            if safe_localname(child) == "name":
                                name = child.text if child.text else "NONAME"
                            if safe_localname(child) == "coordinates":
                                coords = child
                        
                        if coords is not None:
                            coords_text = coords.text.strip()
                            for coord_pair in coords_text.split():
                                try:
                                    lon, lat, *_ = map(float, coord_pair.split(","))
                                    if folder_name.strip().upper() == "FAT":
                                        match = re.search(r'\b([A-Z]\d{1,2})\b', name, re.IGNORECASE)
                                        fat_id = match.group(1).upper() if match else name.strip().upper()
                                        placemarks_fat.append({"fat_id": fat_id, "lat": lat, "lon": lon})
                                    elif "POLE" in folder_name.upper():
                                        placemarks_pole.append({"name": name.strip(), "lat": lat, "lon": lon})
                                    break
                                except:
                                    continue
        
        return placemarks_fat, placemarks_pole
    
    def build_fat_to_pole_map(self, max_distance: float = 15) -> Dict[str, Dict]:
        """Build mapping of FAT to nearest POLE."""
        fat_to_pole = {}
        for fat in self.fats:
            fat_id, flat, flon = fat["fat_id"], fat["lat"], fat["lon"]
            best_match, best_distance = None, float("inf")
            for pole in self.poles:
                d = haversine(flat, flon, pole["lat"], pole["lon"])
                if d < best_distance:
                    best_distance = d
                    best_match = pole
            if best_match and best_distance <= max_distance:
                fat_to_pole[fat_id] = best_match
        return fat_to_pole
    
    def build_fat_to_hpc_map(self, max_distance: float = 25) -> Dict[str, List]:
        """Build mapping of FAT to nearest HP Cover."""
        fat_to_hpc = {}
        for fat in self.fats:
            fat_id, flat, flon = fat["fat_id"], fat["lat"], fat["lon"]
            best_match, best_distance = None, float("inf")
            for hp in self.hp_points:
                d = haversine(flat, flon, float(hp[1]), float(hp[2]))
                if d < best_distance:
                    best_distance = d
                    best_match = hp
            if best_match and best_distance <= max_distance:
                fat_to_hpc[fat_id] = best_match
        return fat_to_hpc
    
    def find_hp_cover_folders(self) -> List:
        """Find all HP COVER folders."""
        hp_cover_folders = []
        for folder in self.root.iter():
            if safe_localname(folder) == "Folder":
                name_elem = None
                for child in folder:
                    if safe_localname(child) == "name":
                        name_elem = child
                        break
                if name_elem is not None and name_elem.text and "HP" in name_elem.text.upper() and "COVER" in name_elem.text.upper() and "UNCOVER" not in name_elem.text.upper():
                    hp_cover_folders.append(folder)
        return hp_cover_folders
    
    def extract_points_recursive(self, element, current_folder: str = "") -> List[List]:
        """Recursively extract points from folders."""
        points = []
        for child in element:
            tag = safe_localname(child)
            if tag == "Folder":
                name_elem = None
                for sub in child:
                    if safe_localname(sub) == "name":
                        name_elem = sub
                        break
                subfolder_name = name_elem.text.strip() if name_elem is not None else current_folder
                new_folder = f"{current_folder}/{subfolder_name}" if current_folder else subfolder_name
                points.extend(self.extract_points_recursive(child, new_folder))
            elif tag == "Placemark":
                name = "No Name"
                lon, lat = "", ""
                for sub in child.iter():
                    if safe_localname(sub) == "name":
                        name = sub.text.strip() if sub.text else "No Name"
                    if safe_localname(sub) == "coordinates":
                        lon, lat = sub.text.strip().split(",")[:2]
                if lat and lon:
                    points.append([name, lat, lon, current_folder])
        return points
    
    def clean_region_name(self, text: str) -> str:
        """Clean region name by removing common prefixes."""
        if not text:
            return ""
        text = str(text)
        remove_words = [
            "Provinsi ", "Kecamatan ", "Kabupaten ", "Kab. ", "Kota ",
            "Desa ", "Kelurahan "
        ]
        for rw in remove_words:
            text = text.replace(rw, "")
        return text.strip().upper()
    
    def reverse_geocode(self, lat: float, lon: float) -> Dict[str, str]:
        """Reverse geocode coordinates using Nominatim."""
        try:
            key = f"{round(float(lat), 6)},{round(float(lon), 6)}"
            if key in self.geocode_cache:
                return self.geocode_cache[key]
            
            url = "https://nominatim.openstreetmap.org/reverse"
            headers = {"User-Agent": "Mozilla/5.0 APD_HPDB_API"}
            params = {
                "lat": lat, "lon": lon,
                "format": "jsonv2",
                "addressdetails": 1,
                "zoom": 18
            }
            
            response = self.session.get(url, params=params, headers=headers, timeout=30)
            time.sleep(1.2)  # Respect rate limiting
            
            if response.status_code != 200:
                return self._empty_geo_result()
            
            data = response.json()
            addr = data.get("address", {})
            
            # Extract street/road name
            street_name = (
                addr.get("road") or
                addr.get("street") or
                addr.get("path") or
                addr.get("pedestrian") or
                addr.get("track") or
                addr.get("highway") or
                addr.get("locality") or
                ""
            )
            
            result = {
                "province": self.clean_region_name(addr.get("state", "")),
                "kabupaten": self.clean_region_name(
                    addr.get("city") or addr.get("county") or 
                    addr.get("municipality") or addr.get("state_district", "")
                ),
                "kecamatan": self.clean_region_name(
                    addr.get("subdistrict") or addr.get("district") or
                    addr.get("city_district") or addr.get("suburb") or
                    addr.get("borough") or addr.get("quarter") or
                    addr.get("neighbourhood", "")
                ),
                "desa": self.clean_region_name(
                    addr.get("village") or addr.get("hamlet") or
                    addr.get("neighbourhood") or addr.get("quarter", "")
                ),
                "kodepos": str(addr.get("postcode", "")).strip(),
                "jalan": str(street_name).strip().upper()
            }
            
            self.geocode_cache[key] = result
            return result
        
        except Exception:
            return self._empty_geo_result()
    
    def _empty_geo_result(self) -> Dict[str, str]:
        """Return empty geocode result."""
        return {"province": "", "kabupaten": "", "kecamatan": "", "desa": "", "kodepos": "", "jalan": ""}
    
    def get_fdt_coords(self) -> Tuple[Optional[str], Optional[str]]:
        """Get FDT coordinates from KML."""
        try:
            for folder in self.root.iter():
                if safe_localname(folder) == "Folder":
                    name_elem = None
                    for child in folder:
                        if safe_localname(child) == "name":
                            name_elem = child
                            break
                    if name_elem is not None and name_elem.text and name_elem.text.strip().upper() == "FDT":
                        for pm in folder:
                            if safe_localname(pm) == "Placemark":
                                for coord in pm.iter():
                                    if safe_localname(coord) == "coordinates":
                                        lon, lat = map(float, coord.text.strip().split(",")[:2])
                                        return f"{lat:.5f}", f"{lon:.5f}"
        except Exception:
            pass
        return None, None

    def get_all_fdt_coords(self) -> Dict[str, Tuple[str, str]]:
        """Get coordinates for all FDTs, mapped by FDT name (e.g. 'FDT 01' -> (lat, lon))"""
        fdt_coords = {}
        try:
            for folder in self.root.iter():
                if safe_localname(folder) == "Folder":
                    name_elem = None
                    for child in folder:
                        if safe_localname(child) == "name":
                            name_elem = child
                            break
                    if name_elem is not None and name_elem.text and "FDT" in name_elem.text.strip().upper():
                        for pm in folder:
                            if safe_localname(pm) == "Placemark":
                                pm_name_el = pm.find(".//name")
                                pm_name = pm_name_el.text.strip().upper() if pm_name_el is not None else ""
                                if "FDT" in pm_name:
                                    fdt_match = re.search(r'\bFDT\s*(\d+)\b', pm_name, re.IGNORECASE)
                                    if fdt_match:
                                        fdt_key = f"FDT {int(fdt_match.group(1)):02d}"
                                    else:
                                        fdt_key = "FDT 01"
                                        
                                    for coord in pm.iter():
                                        if safe_localname(coord) == "coordinates":
                                            parts = coord.text.strip().split(",")
                                            if len(parts) >= 2:
                                                lon, lat = map(float, parts[:2])
                                                fdt_coords[fdt_key] = (f"{lat:.5f}", f"{lon:.5f}")
                                            break
        except Exception as e:
            print("Error in get_all_fdt_coords:", e)
        return fdt_coords
    
    def process(self) -> Dict[str, Any]:
        """Main processing method."""
        if not self.root:
            return {"status": "error", "message": "No KML/KMZ loaded"}
        
        # Parse FAT and POLE
        self.fats, self.poles = self.parse_fat_pole()
        fat_to_pole = self.build_fat_to_pole_map(max_distance=15)
        
        # Parse HP Cover points
        hp_folders = self.find_hp_cover_folders()
        self.hp_points = []
        for folder in hp_folders:
            parent = folder.getparent()
            parent_name = ""
            if parent is not None:
                for child in parent:
                    if safe_localname(child) == "name":
                        parent_name = child.text.strip() if child.text else ""
                        break
            self.hp_points.extend(self.extract_points_recursive(folder, parent_name))
        
        fat_to_hpc = self.build_fat_to_hpc_map(max_distance=25)
        
        # Group hp_points by FDT
        hp_points_by_fdt = defaultdict(list)
        for row in self.hp_points:
            name, lat, lon, folder_path = row
            fdt_name = get_fdt_name_from_folder_path(folder_path)
            hp_points_by_fdt[fdt_name].append(row)
            
        sorted_fdts = sorted(hp_points_by_fdt.keys())
        if not sorted_fdts:
            sorted_fdts = ["FDT 01"] # Fallback
            
        # Process rows per FDT
        fdt_processed_rows = {}
        skip_numbers = {11, 12, 23, 24, 35, 36, 47, 48}
        
        for fdt_name in sorted_fdts:
            rows_in_fdt = hp_points_by_fdt.get(fdt_name, [])
            
            # Calculate FAT ID statistics for this FDT
            fat_id_nums = {}
            for row in rows_in_fdt:
                name, lat, lon, folder_path = row
                match = re.search(r'\b([A-Z]\d{1,2})\b', folder_path + " " + name, re.IGNORECASE)
                fat_id = match.group(1).upper() if match else ""
                if not fat_id:
                    continue
                prefix = fat_id[0].upper()
                num = int(re.findall(r'\d+', fat_id)[0]) if re.findall(r'\d+', fat_id) else 0
                if prefix not in fat_id_nums:
                    fat_id_nums[prefix] = []
                fat_id_nums[prefix].append(num)
            fat_id_max = {k: max(v) for k, v in fat_id_nums.items()}
            
            # Process rows
            fat_port_counters = {}
            fdt_port_counter = 1
            core_number = 1
            last_fat_id = ""
            last_prefix = ""
            
            processed_list = []
            
            for row in rows_in_fdt:
                name, lat, lon, folder_path = row
                name_parts = [p.strip() for p in re.split(r"[,/ ]", name) if p.strip()]
                name_1 = name_parts[0] if len(name_parts) > 0 else ""
                name_2 = name_parts[1] if len(name_parts) > 1 else ""
                
                match = re.search(r'\b([A-Z]\d{1,2})\b', folder_path + " " + name, re.IGNORECASE)
                fat_id = match.group(1).upper() if match else ""
                if not fat_id:
                    continue
                
                prefix = fat_id[0].upper()
                num_part = int(re.findall(r'\d+', fat_id)[0]) if re.findall(r'\d+', fat_id) else 0
                
                if fat_id != last_fat_id:
                    fat_port_counters[fat_id] = 1
                    last_fat_id = fat_id
                
                if prefix != last_prefix:
                    core_number = 1
                    last_prefix = prefix
                
                fat_port = fat_port_counters[fat_id]
                fat_port_counters[fat_id] += 1
                
                fdt_port = fdt_port_counter if fat_port == 1 else ""
                if fat_port == 1:
                    fdt_port_counter += 1
                
                core = ""
                if fat_port in [1, 2]:
                    while core_number in skip_numbers:
                        core_number += 1
                    core = core_number
                    core_number += 1
                
                line, cap, tube_number = "", "", ""
                if core:
                    line = f"LINE {prefix}"
                    max_num = fat_id_max.get(prefix, 0)
                    if max_num <= 10:
                        cap = "24C/2T"
                    elif max_num <= 15:
                        cap = "36C/3T"
                    elif max_num <= 20:
                        cap = "48C/4T"
                    
                    if num_part <= 5:
                        tube_number = "1"
                    elif num_part <= 10:
                        tube_number = "2"
                    elif num_part <= 15:
                        tube_number = "3"
                    elif num_part <= 20:
                        tube_number = "4"
                
                tray = ""
                if isinstance(fdt_port, int):
                    if 1 <= fdt_port <= 8: tray = 1
                    elif 9 <= fdt_port <= 16: tray = 2
                    elif 17 <= fdt_port <= 24: tray = 3
                    elif 25 <= fdt_port <= 32: tray = 4
                    elif 33 <= fdt_port <= 40: tray = 5
                
                pole_name, pole_lat, pole_lon = "", "", ""
                if fat_id in fat_to_pole:
                    pole_name = fat_to_pole[fat_id]["name"]
                    pole_lat = fat_to_pole[fat_id]["lat"]
                    pole_lon = fat_to_pole[fat_id]["lon"]
                
                hpc_text = ""
                if fat_id in fat_to_hpc:
                    hpc_text = "IN FRONT OF HP NUMBER " + fat_to_hpc[fat_id][0]
                
                processed_list.append([
                    tray, fdt_port, line, cap, tube_number, core, fat_id, fat_port,
                    pole_name, pole_lat, pole_lon, hpc_text,
                    name_1, name_2, lat, lon
                ])
                
            fdt_processed_rows[fdt_name] = processed_list
        
        output_filename = f"{os.path.splitext(self.input_filename)[0]}_with_pole.xlsx"
        
        # If template provided, integrate with template
        if self.apd_template_content:
            return self._integrate_with_template(fdt_processed_rows, output_filename)
        
        # Otherwise, save as raw Excel
        wb = Workbook()
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for idx, fdt_name in enumerate(sorted_fdts):
            if idx == 0:
                ws = wb.active
                ws.title = f"Homepass Database {fdt_name}"
            else:
                ws = wb.create_sheet(title=f"Homepass Database {fdt_name}")
                
            headers = [
                "FDT Tray (Front)", "FDT Port", "Line", "Capacity",
                "Tube Colour", "Core Number", "FAT ID",
                "FAT Port", "POLE_Name", "POLE_Lat", "POLE_Lon",
                "HP_Cover", "name_1", "name_2", "latitude", "longitude"
            ]
            ws.append(headers)
            
            for row in fdt_processed_rows.get(fdt_name, []):
                ws.append(row)
                
            for row in ws.iter_rows(min_row=2):
                for cell_idx, cell in enumerate(row, 1):
                    if cell.value not in (None, ""):
                        cell.border = thin_border
                    if cell_idx == 5:
                        if cell.value == "1":
                            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                        elif cell.value == "2":
                            cell.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
                        elif cell.value == "3":
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif cell.value == "4":
                            cell.fill = PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid")
                            
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        return {
            "status": "success",
            "filename": output_filename,
            "content": output_buffer.getvalue()
        }
    
    def _integrate_with_template(self, fdt_processed_rows: Dict[str, List], output_filename: str) -> Dict[str, Any]:
        """Integrate data with APD template."""
        wb_template = load_workbook(io.BytesIO(self.apd_template_content))
        
        headers = [
            "FDT Tray (Front)", "FDT Port", "Line", "Capacity",
            "Tube Colour", "Core Number", "FAT ID",
            "FAT Port", "POLE_Name", "POLE_Lat", "POLE_Lon",
            "HP_Cover", "name_1", "name_2", "latitude", "longitude"
        ]
        
        header_map = {
            "FDT Tray (Front)": "A",
            "FDT Port": "B",
            "Line": "C",
            "Capacity": "D",
            "Tube Colour": "E",
            "Core Number": "F",
            "FAT ID": ["G", "AT"],
            "FAT Port": "H",
            "POLE_Name": "I",
            "POLE_Lat": "J",
            "POLE_Lon": "K",
            "HP_Cover": "L",
            "name_1": "AJ",
            "name_2": "AJ",
            "latitude": "AV",
            "longitude": "AW",
        }
        
        tube_colors = {
            "1": "ADD8E6",
            "2": "FFD580",
            "3": "90EE90",
            "4": "D2B48C",
        }
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Get FDT coordinates mapped by FDT name
        all_fdt_coords = self.get_all_fdt_coords()
        
        # Clean and set location name (remove dash)
        file_base_name = os.path.splitext(os.path.basename(self.input_filename))[0]
        file_base_name_clean = file_base_name.replace("APD_HPDB_", "")
        parts = file_base_name_clean.split(" ", 1)
        lokasi_nama = parts[1].strip() if len(parts) > 1 else file_base_name_clean.strip()
        lokasi_nama_clean = lokasi_nama.replace("-", " ")
        
        sorted_fdts = sorted(fdt_processed_rows.keys())
        
        # 1. Create/prepare worksheets for each FDT in sorted order
        fdt_worksheets = {}
        template_ws = wb_template.worksheets[0]
        
        if len(sorted_fdts) <= 1:
            fdt_name = sorted_fdts[0] if sorted_fdts else "FDT 01"
            fdt_worksheets[fdt_name] = template_ws
        else:
            for idx, fdt_name in enumerate(sorted_fdts):
                if idx == 0:
                    ws = template_ws
                    ws.title = f"Homepass Database {fdt_name}"
                else:
                    ws = wb_template.copy_worksheet(template_ws)
                    ws.title = f"Homepass Database {fdt_name}"
                fdt_worksheets[fdt_name] = ws
        
        # 2. Populate each worksheet with its own FDT data
        for fdt_name in sorted_fdts:
            ws = fdt_worksheets[fdt_name]
            data_rows = fdt_processed_rows.get(fdt_name, [])
            
            start_row = 10
            for i, row in enumerate(data_rows, start=start_row):
                row_dict = dict(zip(headers, row))
                
                for key, target in header_map.items():
                    if key not in row_dict:
                        continue
                    val = row_dict[key]
                    if val in (None, ""):
                        continue
                    
                    if isinstance(target, list):
                        for col in target:
                            cell = ws[f"{col}{i}"]
                            cell.value = val
                            cell.border = thin_border
                    elif key == "name_1":
                        combined = str(val)
                        if row_dict.get("name_2"):
                            combined += " " + str(row_dict["name_2"])
                        cell = ws[f"{target}{i}"]
                        cell.value = combined
                        cell.border = thin_border
                    elif key == "name_2":
                        continue
                    else:
                        cell = ws[f"{target}{i}"]
                        if key in ["POLE_Lat", "POLE_Lon", "latitude", "longitude"]:
                            try:
                                val = float(val)
                            except:
                                pass
                        
                        cell.value = val
                        cell.border = thin_border
                        
                        if key == "Tube Colour":
                            color_hex = tube_colors.get(str(val).strip(), None)
                            if color_hex:
                                cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            # Set FDT coordinates for this sheet
            lat_fdt, lon_fdt = all_fdt_coords.get(fdt_name, (None, None))
            # Fallback if specific not found, use first found or None
            if not lat_fdt and all_fdt_coords:
                lat_fdt, lon_fdt = next(iter(all_fdt_coords.values()))
            if lat_fdt and lon_fdt:
                ws["N6"].value = f": {lat_fdt},{lon_fdt}"
            
            # Set cluster/location name
            ws["C5"].value = lokasi_nama_clean
            ws["Y10"].value = lokasi_nama_clean
            ws["Z10"].value = lokasi_nama_clean
            
            # Reverse geocode and autocopy (using the first populated row, i=10)
            max_row_data = start_row + len(data_rows) - 1
            if len(data_rows) > 0:
                lat_source = ws["J10"].value
                lon_source = ws["K10"].value
                
                if lat_source in (None, "") or lon_source in (None, ""):
                    lat_source = ws["AV10"].value
                    lon_source = ws["AW10"].value
                
                geo_result = self._empty_geo_result()
                if lat_source not in (None, "") and lon_source not in (None, ""):
                    geo_result = self.reverse_geocode(lat_source, lon_source)
                
                ws["M10"] = geo_result["province"]
                ws["N10"] = geo_result["kabupaten"]
                ws["W10"] = geo_result["kabupaten"]
                ws["O10"] = geo_result["kecamatan"]
                ws["P10"] = geo_result["desa"]
                ws["Q10"] = geo_result["kodepos"]
                ws["AD10"] = geo_result["jalan"]  # Nama jalan
                
                # Autocopy geocoding and cluster data to all rows
                for r in range(11, max_row_data + 1):
                    ws[f"M{r}"] = ws["M10"].value
                    ws[f"N{r}"] = ws["N10"].value
                    ws[f"W{r}"] = ws["W10"].value
                    ws[f"O{r}"] = ws["O10"].value
                    ws[f"P{r}"] = ws["P10"].value
                    ws[f"Q{r}"] = ws["Q10"].value
                    ws[f"AD{r}"] = ws["AD10"].value  # Autocopy nama jalan
                    ws[f"Y{r}"] = ws["Y10"].value    # Autocopy cluster Y
                    ws[f"Z{r}"] = ws["Z10"].value    # Autocopy cluster Z
            
            # Count unique FAT IDs to determine capacity (Cell N3)
            fat_id_col_idx = headers.index("FAT ID") if "FAT ID" in headers else -1
            unique_fat_ids = set()
            if fat_id_col_idx >= 0:
                for row in data_rows:
                    fat_id = row[fat_id_col_idx]
                    if fat_id and str(fat_id).strip():
                        unique_fat_ids.add(str(fat_id).strip().upper())
            
            fat_count = len(unique_fat_ids)
            if fat_count <= 20:
                capacity = "48C"
            elif fat_count <= 30:
                capacity = "72C"
            elif fat_count <= 40:
                capacity = "96C"
            else:
                capacity = "96C"  # Default untuk lebih dari 40
            
            ws["N3"] = capacity
            
            # Set formula for N4
            if len(data_rows) > 0:
                ws["N4"] = f'=": "&COUNTA(G10:G{max_row_data})&" HP /Aerial"'
            else:
                ws["N4"] = '=": 0 HP /Aerial"'
            
            # Delete unused rows from template
            last_template_row = ws.max_row
            if max_row_data < last_template_row:
                delete_start = max_row_data + 1
                delete_count = last_template_row - max_row_data
                ws.delete_rows(delete_start, delete_count)
        
        # Save final output
        output_buffer = io.BytesIO()
        wb_template.save(output_buffer)
        output_buffer.seek(0)
        
        # Generate final filename
        final_name = self._generate_final_filename(output_filename)
        
        return {
            "status": "success",
            "filename": final_name,
            "content": output_buffer.getvalue()
        }
    
    def _generate_final_filename(self, output_filename: str) -> str:
        """Generate final filename according to format."""
        excel_name = os.path.basename(output_filename)
        excel_base = os.path.splitext(excel_name)[0]
        
        excel_base = excel_base.replace("_with_pole", "")
        
        # Remove old prefixes if they exist in the original filename
        for old_prefix in ["APD_HPDB_", "HPDB_"]:
            if excel_base.startswith(old_prefix):
                excel_base = excel_base[len(old_prefix):]
        
        # New prefix format requested by user
        prefix = "HPDB_-"
        
        parts = excel_base.split(" ", 1)
        lokasi_nama_final = parts[1].strip() if len(parts) > 1 else excel_base.strip()
        
        # Ensure there's a space after the prefix
        return f"{prefix} {lokasi_nama_final}.xlsx"


def process_apd_hpdb(
    kml_content: bytes,
    filename: str,
    apd_template_content: bytes = None
) -> Dict[str, Any]:
    """
    Process APD HPDB from KML or KMZ file.
    
    Args:
        kml_content: Raw bytes of KML or KMZ file
        filename: Original filename
        apd_template_content: Optional APD template Excel bytes
    
    Returns:
        Dict with status, filename, and content bytes
    """
    is_kmz = filename.lower().endswith(".kmz")
    engine = APDEngine(apd_template_content=apd_template_content)
    engine.load_kml(kml_content, filename, is_kmz=is_kmz)
    return engine.process()
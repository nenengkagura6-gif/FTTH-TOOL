"""
Engine for processing KML/KMZ files and generating BOQ Excel
"""
import io
import os
import re
from typing import Dict, List, Tuple, Any
from geopy.distance import geodesic
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from xml.dom import minidom
from collections import defaultdict

from utils.commons import (
    parse_kml_content, find_all_folders, get_folder_name,
    parse_coords, clean_project_name
)


def get_direct_child_name(node: minidom.Element) -> str:
    """Get the text value of the direct child <name> tag of a node."""
    for child in node.childNodes:
        if child.nodeType == minidom.Node.ELEMENT_NODE and child.nodeName == "name":
            if child.firstChild:
                return child.firstChild.nodeValue.strip()
    return ""


def get_fdt_name_from_ancestors(node: minidom.Element) -> str:
    """Find the FDT name by traversing up the parent nodes to find a Folder containing FDT name."""
    curr = node
    while curr is not None:
        if curr.nodeType == minidom.Node.ELEMENT_NODE and curr.nodeName == "Folder":
            name = get_direct_child_name(curr)
            match = re.search(r'\bFDT\s*(\d+)\b', name, re.IGNORECASE)
            if match:
                return f"FDT {int(match.group(1)):02d}"
        curr = curr.parentNode
    return "FDT 01" # Default fallback


class KMLEngine:
    """Engine to process KML/KMZ files and generate BOQ Excel."""
    
    def __init__(self, template_content: bytes = None):
        if template_content:
            self.template_content = template_content
        else:
            base_dir = Path(__file__).resolve().parent.parent
            default_template = base_dir / "templates" / "default_boq.xlsx"
            
            if default_template.exists():
                with open(default_template, "rb") as f:
                    self.template_content = f.read()
            else:
                self.template_content = None
        
        self.doc = None
        self.wb = None
        self.sheet_ae = None
        self.sheet_bo = None
        self.input_filename = ""
    
    def load_kml(self, content: bytes, filename: str, is_kmz: bool = False) -> Dict[str, Any]:
        """Load KML/KMZ content."""
        self.input_filename = filename
        self.doc = parse_kml_content(content, is_kmz)
        return {"status": "success", "filename": filename}
    
    def load_template(self, template_content: bytes) -> Dict[str, Any]:
        """Load Excel template from bytes."""
        self.template_content = template_content
        return {"status": "success"}
    
    def _init_workbook(self) -> None:
        """Initialize workbook from template."""
        if self.template_content:
            self.wb = load_workbook(io.BytesIO(self.template_content))
        else:
            self.wb = Workbook()
        
        self.sheet_ae = self.wb["BoM AE"] if "BoM AE" in self.wb.sheetnames else self.wb.active
        self.sheet_bo = self.wb["BoQ NRO Cluster"] if "BoQ NRO Cluster" in self.wb.sheetnames else self.wb.active
    
    def _calculate_length_from_placemark(self, placemark: minidom.Element) -> float:
        """Calculate length from a LineString placemark."""
        coords_tags = placemark.getElementsByTagName("coordinates")
        if coords_tags and coords_tags[0].firstChild:
            coord_list = parse_coords(coords_tags[0].firstChild.nodeValue)
            return sum(
                geodesic(coord_list[i], coord_list[i+1]).meters 
                for i in range(len(coord_list)-1)
            )
        return 0
    
    def _safe_add(self, sheet, cell: str, value: float) -> None:
        """Safely add value to a cell."""
        current = sheet[cell].value
        if current is None:
            current_val = 0.0
        elif isinstance(current, (int, float)):
            current_val = float(current)
        else:
            try:
                current_val = float(str(current).strip())
            except ValueError:
                current_val = 0.0
        sheet[cell] = round(current_val + value)
    
    def _is_true_fat_folder(self, name: str) -> bool:
        """Check if folder is a FAT folder (not FAT COVER)."""
        up = (name or "").upper()
        return "FAT" in up and "COVER" not in up
    
    def _count_fat_in_line(self, line_folder: minidom.Element) -> int:
        """Count all Placemarks in FAT subfolders recursively."""
        total = 0
        subfolders = find_all_folders(line_folder)
        for f in subfolders:
            if self._is_true_fat_folder(get_folder_name(f)):
                total += len(f.getElementsByTagName("Placemark"))
        return total
    
    def process(self) -> Dict[str, Any]:
        """Process the KML and generate Excel output."""
        if not self.doc:
            return {"status": "error", "message": "No KML loaded"}
        
        self._init_workbook()
        all_folders = find_all_folders(self.doc.documentElement)
        
        # Detect if KML has multiple FDTs in line folder names
        fdt_map = {"FDT 01": "C", "FDT 02": "I", "FDT 03": "O"}
        detected_fdts = set()
        for folder in all_folders:
            name = get_folder_name(folder).upper()
            m = re.search(r'\bFDT\s*(\d+)\b', name)
            if m:
                detected_fdts.add(f"FDT {int(m.group(1)):02d}")
        is_multi_fdt = len(detected_fdts) > 1
        
        # Process FDT columns
        for fdt_name, col in fdt_map.items():
            self._process_fdt_column(col, all_folders, fdt_name)
        
        # Process FAT per Line
        self._process_fat_per_line(fdt_map, is_multi_fdt)
        
        # Process Pole Counts
        self._process_pole_counts(all_folders, fdt_map, is_multi_fdt)
        
        # Process HP Cover (remains global)
        self._process_hp_cover(all_folders)
        
        # Save output to bytes
        output_buffer = io.BytesIO()
        self.wb.save(output_buffer)
        output_buffer.seek(0)
        
        output_filename = f"Hasil_{os.path.splitext(self.input_filename)[0]}.xlsx"
        
        return {
            "status": "success",
            "filename": output_filename,
            "content": output_buffer.getvalue()
        }
    
    def _process_fdt_column(self, col: str, all_folders: List, target_fdt: str) -> None:
        """Process Distribution Cable for an FDT column."""
        processed_pms = set()  # Track processed Placemarks to avoid double counting
        total_sling_length = 0.0
        
        for sub in all_folders:
            if get_fdt_name_from_ancestors(sub) != target_fdt:
                continue
                
            if "distribution" in get_folder_name(sub).lower():
                for pm in sub.getElementsByTagName("Placemark"):
                    pm_id = id(pm)
                    if pm_id in processed_pms:
                        continue
                    processed_pms.add(pm_id)
                    
                    nm_nodes = pm.getElementsByTagName("name")
                    pm_name = (nm_nodes[0].firstChild.nodeValue if nm_nodes and nm_nodes[0].firstChild else "").upper()
                    length = self._calculate_length_from_placemark(pm)
                    
                    # Process by line and cable type
                    self._process_cable_entry(col, pm_name, length)
            
            # Process Sling Wire
            if "sling" in get_folder_name(sub).lower():
                for pm in sub.getElementsByTagName("Placemark"):
                    pm_id = id(pm)
                    if pm_id in processed_pms:
                        continue
                    processed_pms.add(pm_id)
                    total_sling_length += self._calculate_length_from_placemark(pm)
                    
        if total_sling_length > 0:
            self.sheet_ae[f"{col}15"] = round(total_sling_length)
    
    def _process_cable_entry(self, col: str, pm_name: str, length: float) -> None:
        """Process a cable entry based on line and type."""
        lines = ["LINE A", "LINE B", "LINE C", "LINE D"]
        cable_types = {"24C": 0, "36C": 4, "48C": 8}
        
        for line_idx, line in enumerate(lines):
            if line in pm_name:
                for cable, offset in cable_types.items():
                    if cable in pm_name:
                        cell = f"{col}{2 + line_idx + offset}"
                        self._safe_add(self.sheet_ae, cell, length)
                break
    
    def _process_fat_per_line(self, fdt_map: Dict[str, str], is_multi_fdt: bool) -> None:
        """Process FAT counts per Line, split per FDT (or all columns for single FDT)."""
        for line_folder in self.doc.getElementsByTagName("Folder"):
            line_name = get_folder_name(line_folder).upper()
            
            matched_line = None
            for l in ["LINE A", "LINE B", "LINE C", "LINE D"]:
                if l in line_name:
                    matched_line = l
                    break
            
            if matched_line:
                total_fat = 0
                has_fat = False
                for sub_folder in line_folder.getElementsByTagName("Folder"):
                    sub_name = get_folder_name(sub_folder).upper()
                    if sub_name == "FAT":
                        total_fat = len(sub_folder.getElementsByTagName("Placemark"))
                        has_fat = True
                        break
                
                if has_fat:
                    row_map = {"LINE A": 36, "LINE B": 37, "LINE C": 38, "LINE D": 39}
                    row = row_map[matched_line]
                    if is_multi_fdt:
                        # Multi-FDT: write only to the specific FDT column
                        fdt_name = get_fdt_name_from_ancestors(line_folder)
                        col = fdt_map.get(fdt_name)
                        if col:
                            self.sheet_ae[f"{col}{row}"] = total_fat
                    else:
                        # Single FDT: write to ALL columns (backward compatible)
                        for col in fdt_map.values():
                            self.sheet_ae[f"{col}{row}"] = total_fat
    
    def _process_pole_counts(self, all_folders: List, fdt_map: Dict[str, str], is_multi_fdt: bool) -> None:
        """Process pole counts, split per FDT or all columns for single FDT."""
        target_rows = {
            "new pole 7-4": 54,
            "new pole 7-3": 55,
            "new pole 7-2.5": 56,
            "new pole 9-4": 58,
            "existing pole emr 7-4": 61,
        }
        
        processed_pms = defaultdict(set)
        pole_totals = defaultdict(int)  # key: (fdt_name, target_name)
        
        for folder in all_folders:
            folder_name = get_folder_name(folder).strip().lower()
            for target_name, row in target_rows.items():
                if folder_name == target_name:
                    fdt_name = get_fdt_name_from_ancestors(folder)
                    for pm in folder.getElementsByTagName("Placemark"):
                        pm_id = id(pm)
                        key = (fdt_name, target_name)
                        if pm_id not in processed_pms[key]:
                            processed_pms[key].add(pm_id)
                            pole_totals[key] += 1
                            
        for (fdt_name, target_name), total in pole_totals.items():
            row = target_rows[target_name]
            if is_multi_fdt:
                # Multi-FDT: write only to the specific FDT column
                col = fdt_map.get(fdt_name)
                if col:
                    self.sheet_ae[f"{col}{row}"] = total
            else:
                # Single FDT: write to ALL columns (backward compatible)
                for col in fdt_map.values():
                    self.sheet_ae[f"{col}{row}"] = total
    
    def _process_hp_cover(self, all_folders: List) -> None:
        """Process HP Cover count."""
        hp_cover_total = 0
        processed_pms = set()
        for folder in all_folders:
            if "hp cover" in get_folder_name(folder).lower():
                for pm in folder.getElementsByTagName("Placemark"):
                    pm_id = id(pm)
                    if pm_id not in processed_pms:
                        processed_pms.add(pm_id)
                        hp_cover_total += 1
        
        self.sheet_bo["O5"] = hp_cover_total
        self.sheet_bo["O3"] = clean_project_name(self.input_filename)


def process_kml_to_excel(
    kml_content: bytes,
    filename: str,
    template_content: bytes = None,
    is_kmz: bool = False
) -> Dict[str, Any]:
    """
    Process KML/KMZ file and generate BOQ Excel.
    
    Args:
        kml_content: Raw bytes of KML/KMZ file
        filename: Original filename
        template_content: Optional Excel template bytes
        is_kmz: Whether the file is KMZ format
    
    Returns:
        Dict with status, filename, and content bytes
    """
    engine = KMLEngine(template_content)
    engine.load_kml(kml_content, filename, is_kmz)
    return engine.process()
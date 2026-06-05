"""
Engine for extracting layers and folders from KML/KMZ files and generating summarized Excel reports.
"""
import io
import os
from typing import Dict, List, Tuple, Any
from geopy.distance import geodesic
from xml.dom import minidom
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.commons import (
    parse_kml_content, get_folder_name, parse_coords, clean_project_name
)


class KMLExtractorEngine:
    """Engine to extract folder and element summaries from KML/KMZ files."""
    
    def __init__(self):
        self.doc = None
        self.filename = ""
        self.is_kmz = False

    def load_kml(self, content: bytes, filename: str, is_kmz: bool = False) -> None:
        """Load and parse the KML/KMZ content."""
        self.filename = filename
        self.is_kmz = is_kmz
        self.doc = parse_kml_content(content, is_kmz)

    def _get_folder_path(self, node: minidom.Element) -> List[str]:
        """Traverse upwards to find the full folder hierarchy path."""
        path = []
        curr = node.parentNode
        while curr:
            if curr.nodeType == curr.ELEMENT_NODE and curr.tagName == "Folder":
                name = get_folder_name(curr)
                if name:
                    path.insert(0, name)
            curr = curr.parentNode
        return path

    def _determine_layer_folder(self, path: List[str]) -> str:
        """Determine the main top-level folder, skipping the wrapper folder if present."""
        if not path:
            return "Tanpa Folder"
        if len(path) == 1:
            return path[0]
        
        # Clean filename to see if the first folder is just a filename wrapper
        clean_file = clean_project_name(self.filename).lower()
        first_folder = path[0].lower()
        
        # If the first folder name matches or contains the filename (or vice versa), skip it
        if clean_file == first_folder or clean_file in first_folder or first_folder in clean_file:
            return path[1]
        return path[0]

    def _calculate_line_length(self, pm: minidom.Element) -> float:
        """Calculate length of a LineString in meters."""
        coords_tags = pm.getElementsByTagName("coordinates")
        if coords_tags and coords_tags[0].firstChild:
            coord_val = coords_tags[0].firstChild.nodeValue
            coord_list = parse_coords(coord_val)
            if len(coord_list) >= 2:
                return sum(
                    geodesic(coord_list[i], coord_list[i+1]).meters 
                    for i in range(len(coord_list)-1)
                )
        return 0.0

    def _get_placemark_name(self, pm: minidom.Element) -> str:
        """Get the name tag value of a placemark."""
        names = pm.getElementsByTagName("name")
        if names and names[0].firstChild:
            return names[0].firstChild.nodeValue.strip()
        return ""

    def process(self) -> Dict[str, Any]:
        """Process KML/KMZ and generate stylized Excel report."""
        if not self.doc:
            return {"status": "error", "message": "No KML loaded"}

        # Groups structure: layer_folder -> cable_type -> statistics
        # cable_type is empty string for non-cable folders
        groups = {}

        all_placemarks = self.doc.getElementsByTagName("Placemark")
        for pm in all_placemarks:
            path = self._get_folder_path(pm)
            layer = self._determine_layer_folder(path)
            
            # Identify geometry type
            is_point = len(pm.getElementsByTagName("Point")) > 0
            is_line = len(pm.getElementsByTagName("LineString")) > 0
            is_poly = len(pm.getElementsByTagName("Polygon")) > 0
            
            # Only count valid types
            if not (is_point or is_line or is_poly):
                continue

            # Determine if folder represents cable/kabel
            is_cable = "cable" in layer.lower() or "kabel" in layer.lower()
            
            if is_cable and is_line:
                pm_name = self._get_placemark_name(pm)
                cable_type = pm_name if pm_name else "Kabel Tanpa Nama"
            else:
                cable_type = ""

            if layer not in groups:
                groups[layer] = {}
            if cable_type not in groups[layer]:
                groups[layer][cable_type] = {
                    "points": 0,
                    "polygons": 0,
                    "lines": 0,
                    "length_m": 0.0
                }

            stats = groups[layer][cable_type]
            if is_point:
                stats["points"] += 1
            elif is_poly:
                stats["polygons"] += 1
            elif is_line:
                stats["lines"] += 1
                stats["length_m"] += self._calculate_line_length(pm)

        # Generate openpyxl Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ringkasan Ekstraksi"

        # Show Gridlines
        ws.views.sheetView[0].showGridLines = True

        # Styles definition
        font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Segoe UI", size=10, color="000000")
        font_total = Font(name="Segoe UI", size=11, bold=True, color="000000")
        
        fill_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
        fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Dark Grayish Blue
        fill_zebra = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid") # Light Zebra
        fill_total = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid") # Gray

        thin_side = Side(border_style="thin", color="D5D8DC")
        double_side = Side(border_style="double", color="2C3E50")
        thick_top = Side(border_style="medium", color="2C3E50")
        
        border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(top=thick_top, bottom=double_side)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # 1. Title Block
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"LAPORAN EKSTRAKSI FOLDER KML: {clean_project_name(self.filename)}"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = align_center
        ws.row_dimensions[1].height = 40

        # Subtitle
        ws.merge_cells("A2:H2")
        sub_cell = ws["A2"]
        sub_cell.value = "Dibuat secara otomatis oleh FTTH Tool Platform"
        sub_cell.font = Font(name="Segoe UI", size=10, italic=True, color="566573")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # Leave row 3 empty
        ws.row_dimensions[3].height = 10

        # 2. Table Headers
        headers = [
            "No",
            "Nama Folder",
            "Tipe Kabel / Detail",
            "Titik (Points)",
            "Area (Polygons)",
            "Rute (Lines)",
            "Panjang Total (m)",
            "Panjang Total (km)"
        ]

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = h
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = Border(top=thin_side, bottom=Side(border_style="medium", color="FFFFFF"), left=thin_side, right=thin_side)
        ws.row_dimensions[4].height = 25

        # 3. Table Body
        row_num = 5
        no_counter = 1

        # Sort layers alphabetically for consistent output
        sorted_layers = sorted(groups.keys())

        total_points = 0
        total_polygons = 0
        total_lines = 0
        total_length_m = 0.0

        for layer in sorted_layers:
            # Sort cable types, putting empty string (non-cables or standard stats) first
            sorted_cables = sorted(groups[layer].keys())
            
            for cable in sorted_cables:
                stats = groups[layer][cable]
                
                # Write rows
                cells = [
                    ws.cell(row=row_num, column=1, value=no_counter),
                    ws.cell(row=row_num, column=2, value=layer),
                    ws.cell(row=row_num, column=3, value=cable if cable else "-"),
                    ws.cell(row=row_num, column=4, value=stats["points"]),
                    ws.cell(row=row_num, column=5, value=stats["polygons"]),
                    ws.cell(row=row_num, column=6, value=stats["lines"]),
                    ws.cell(row=row_num, column=7, value=round(stats["length_m"])),
                    ws.cell(row=row_num, column=8, value=round(stats["length_m"] / 1000.0, 3))
                ]

                # Update totals
                total_points += stats["points"]
                total_polygons += stats["polygons"]
                total_lines += stats["lines"]
                total_length_m += stats["length_m"]

                # Apply formatting
                is_even = (no_counter % 2 == 0)
                for col_idx, cell in enumerate(cells, 1):
                    cell.font = font_body
                    cell.border = border_cell
                    if is_even:
                        cell.fill = fill_zebra
                    
                    # Alignments
                    if col_idx in [1, 4, 5, 6]:
                        cell.alignment = align_center
                    elif col_idx in [2, 3]:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_right

                    # Number formats
                    if col_idx in [4, 5, 6]:
                        cell.number_format = "#,##0"
                    elif col_idx == 7:
                        cell.number_format = "#,##0"
                    elif col_idx == 8:
                        cell.number_format = "#,##0.000"

                ws.row_dimensions[row_num].height = 20
                row_num += 1
                no_counter += 1

        # 4. Total Summary Row
        total_cells = [
            ws.cell(row=row_num, column=1, value=""),
            ws.cell(row=row_num, column=2, value="TOTAL"),
            ws.cell(row=row_num, column=3, value=""),
            ws.cell(row=row_num, column=4, value=total_points),
            ws.cell(row=row_num, column=5, value=total_polygons),
            ws.cell(row=row_num, column=6, value=total_lines),
            ws.cell(row=row_num, column=7, value=round(total_length_m)),
            ws.cell(row=row_num, column=8, value=round(total_length_m / 1000.0, 3))
        ]

        for col_idx, cell in enumerate(total_cells, 1):
            cell.font = font_total
            cell.fill = fill_total
            cell.border = border_total
            
            # Alignments
            if col_idx == 2:
                cell.alignment = align_left
            elif col_idx in [4, 5, 6]:
                cell.alignment = align_center
                cell.number_format = "#,##0"
            elif col_idx == 7:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            elif col_idx == 8:
                cell.alignment = align_right
                cell.number_format = "#,##0.000"

        ws.row_dimensions[row_num].height = 24

        # 5. Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Avoid using title and subtitle rows for width calculation
            for cell in col[3:]:
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Force column widths for specific columns to look nicer
        ws.column_dimensions['A'].width = 6   # No
        ws.column_dimensions['B'].width = 30  # Nama Folder
        ws.column_dimensions['C'].width = 32  # Tipe Kabel / Detail

        # Save to byte buffer
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        output_filename = f"Ekstraksi_{os.path.splitext(self.filename)[0]}.xlsx"

        return {
            "status": "success",
            "filename": output_filename,
            "content": output_buffer.getvalue(),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }


def process_kml_extractor(
    kml_content: bytes,
    filename: str,
    is_kmz: bool = False
) -> Dict[str, Any]:
    """
    Process KML/KMZ file and extract elements by folder.
    
    Args:
        kml_content: Raw bytes of KML/KMZ file
        filename: Original filename
        is_kmz: Whether the file is KMZ format
    
    Returns:
        Dict with status, filename, and content bytes
    """
    engine = KMLExtractorEngine()
    engine.load_kml(kml_content, filename, is_kmz)
    return engine.process()
